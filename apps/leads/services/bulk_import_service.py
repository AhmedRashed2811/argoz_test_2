"""Bulk lead import from CSV (leads spec §4 bulk_create). Parsing, per-row
validation, duplicate handling and creation all live here so the view/API stays
thin. Distribution follows the lead.bulk_import_distribution policy:
  - AUTO   -> auto-distribute using the company's default method (notifications on)
  - MANUAL -> created requiring manual distribution, no SLA, no notifications
Existing phones are never created; they are reported as duplicates and may be
reactivated (manual distribution, no SLA) on explicit confirmation."""
from __future__ import annotations

import csv
import io
from collections import Counter

from django.core.validators import validate_email
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction
from django.utils import timezone

from apps.accounts.models import Broker, Language
from apps.marketing.constants import ChannelType
from apps.marketing.models import (
    Campaign, EventRecord, ExhibitionRecord, SocialMediaAdRecord,
    StreetAdRecord, TVAdRecord,
)
from apps.policies.constants import PolicyCode
from apps.policies.services import PolicyResolver

from ..constants import ActiveStatus, Origin, SourceCode, StageCode
from ..models import Lead
from ..phone import validate_phone
from .duplicate_service import DuplicateService
from .lead_creation_service import LeadCreationService

DEFAULT_COUNTRY_CODE = "+20"

# Sources accepted in bulk import (interactive sources like walk-in/exhibition
# that require a live salesman pick are intentionally excluded).
_SOURCE_ALIASES = {
    "campaign": SourceCode.CAMPAIGN,
    "broker": SourceCode.BROKER,
    "referral": SourceCode.REFERRAL,
    "call_center": SourceCode.CALL_CENTER,
    "call center": SourceCode.CALL_CENTER,
}

# Campaign channel column -> (ChannelType, record model).
_CHANNEL_COLUMNS = {
    "event": (ChannelType.EVENT, EventRecord),
    "tv_ad": (ChannelType.TV_AD, TVAdRecord),
    "street_ad": (ChannelType.STREET_AD, StreetAdRecord),
    "social_media_ad": (ChannelType.SOCIAL_MEDIA_AD, SocialMediaAdRecord),
    "exhibition": (ChannelType.EXHIBITION, ExhibitionRecord),
}

# Column order for the downloadable template / rejected-rows export.
COLUMNS = [
    "name", "country_code", "phone", "email", "source",
    "broker_name", "campaign_name",
    "event", "tv_ad", "street_ad", "social_media_ad", "exhibition",
    "referrer_name", "language", "notes",
]


class BulkLeadImportService:
    @staticmethod
    def import_csv(*, company, actor, file_bytes: bytes, request_meta=None) -> dict:
        rows = BulkLeadImportService._read(file_bytes)
        if rows is None:
            return {"error": "Could not read the CSV file. Save it as UTF-8 CSV."}

        # Broker self-import: if the uploader is themselves a broker, every row is
        # a broker lead owned by their brokerage — no source/broker_name columns,
        # just name, country_code, phone, email, language, notes (§8.5).
        from apps.accounts.models import BrokerStatus
        self_broker = Broker.objects.filter(
            company=company, linked_user=actor, status=BrokerStatus.ACTIVE
        ).first()

        method = (PolicyResolver.option_code(
            company, PolicyCode.BULK_IMPORT_DISTRIBUTION, default="MANUAL") or "MANUAL").upper()
        auto = method == "AUTO"

        accepted = 0
        rejected_rows: list[dict] = []
        duplicates: list[dict] = []
        seen_phones: set[str] = set()

        for raw in rows:
            row = {k: (v or "").strip() for k, v in raw.items() if k is not None and isinstance(v, str)}
            phone = _normalize_phone(row.get("phone", ""))
            try:
                clean = BulkLeadImportService._validate_row(
                    company, row, phone, self_broker=self_broker)
            except _RowError as exc:
                rejected_rows.append({**row, "error": str(exc), "error_fields": exc.fields})
                continue

            # Duplicate phone (in DB or earlier in this file) -> never create.
            if phone in seen_phones:
                duplicates.append({**row, "error": "Duplicate phone within this file.",
                                   "error_fields": ["phone"]})
                continue
            dup = DuplicateService.check(company=company, phone=phone)
            if dup.is_duplicate:
                duplicates.append({
                    **row, "phone": phone,
                    "existing_lead_id": str(dup.existing.id),
                    "error": "Phone already exists in the system.",
                    "error_fields": ["phone"],
                })
                continue

            BulkLeadImportService._create(
                company=company, actor=actor, clean=clean,
                auto=auto, request_meta=request_meta)
            accepted += 1
            seen_phones.add(phone)

        reasons = Counter(r["error"] for r in rejected_rows)
        dup_reasons = Counter(d["error"] for d in duplicates)
        return {
            "accepted": accepted,
            "rejected": len(rejected_rows),
            "duplicate_count": len(duplicates),
            "rejected_rows": rejected_rows,
            "duplicates": duplicates,
            "reasons_summary": dict(reasons),
            "duplicate_summary": dict(dup_reasons),
            "distribution": method,
            "columns": COLUMNS,
        }

    @staticmethod
    def build_rejected_xlsx(*, rows: list[dict], columns: list[str]) -> bytes:
        """Build an .xlsx of rejected rows with each value cell filled light red
        for quick scanning (task 9). The header row is left unfilled; a trailing
        'error' column explains each rejection."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        header = list(columns) + ["error"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Rejected leads"
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in rows:
            ws.append([row.get(col, "") if row.get(col) is not None else "" for col in header])
            # Highlight only the value cell(s) the importer flagged for this row;
            # fall back to the whole row if the offending field is unknown.
            fields = set(row.get("error_fields") or [])
            for col_idx, col in enumerate(header, start=1):
                if (fields and col in fields) or (not fields and col != "error"):
                    ws.cell(row=ws.max_row, column=col_idx).fill = red
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    @transaction.atomic
    def reactivate(*, company, actor, phones: list[str], request_meta=None) -> int:
        """Reactivate existing leads matched by phone: active, manual distribution,
        no SLA, no salesman, no notifications (mirrors a broker lead, §8.5)."""
        fresh_stage = StageCode.FRESH
        from ..models import LeadStageDefinition
        fresh = LeadStageDefinition.objects.get(code=fresh_stage)
        count = 0
        wanted = {_normalize_phone(p) for p in phones}
        leads = Lead.objects.filter(company=company, phone__in=wanted)
        for lead in leads:
            lead.active_status = ActiveStatus.ACTIVE
            lead.assigned_salesman = None
            lead.assigned_team = None
            lead.sla_deadline = None
            lead.current_stage = fresh
            lead.last_activity_at = timezone.now()
            lead.save(update_fields=[
                "active_status", "assigned_salesman", "assigned_team",
                "sla_deadline", "current_stage", "last_activity_at", "updated_at"])
            count += 1
        return count

    # ── internals ────────────────────────────────────────────────────────
    @staticmethod
    def _read(file_bytes: bytes):
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            return None
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return None
        # Normalize header names to our lowercase/underscore keys.
        reader.fieldnames = [(_norm_header(h)) for h in reader.fieldnames]
        return list(reader)

    @staticmethod
    def _validate_row(company, row, phone, self_broker=None) -> dict:
        name = row.get("name", "")
        if not name:
            raise _RowError("Name is required.", ["name"])
        if not phone:
            raise _RowError("Phone is required.", ["phone"])
        country = row.get("country_code") or DEFAULT_COUNTRY_CODE
        if not country.startswith("+") or not country[1:].isdigit():
            raise _RowError("Country code must look like +20.", ["country_code"])
        # Country-aware length check (same rules as the Add-Lead form).
        phone_err = validate_phone(country, phone)
        if phone_err:
            raise _RowError(phone_err, ["phone"])
        email = row.get("email", "")
        if email:
            try:
                validate_email(email)
            except DjangoValidationError:
                raise _RowError("Email is not valid.", ["email"])

        language = _resolve_language(row.get("language", ""), company)

        # Broker self-import: simplified template, lead owned by the uploader's
        # brokerage; no source/broker_name/campaign columns required.
        if self_broker is not None:
            return dict(
                source=SourceCode.BROKER, name=name, phone=phone,
                country_code=country, email=email, language=language,
                notes=row.get("notes", ""), broker=self_broker,
            )

        source = _SOURCE_ALIASES.get(row.get("source", "").lower())
        if source is None:
            raise _RowError(
                "Source must be one of: campaign, broker, referral, call_center.",
                ["source"])

        clean = dict(
            source=source, name=name, phone=phone, country_code=country,
            email=email, language=language, notes=row.get("notes", ""),
        )

        if source == SourceCode.BROKER:
            broker_name = row.get("broker_name", "")
            if not broker_name:
                raise _RowError("Broker name is required for broker leads.", ["broker_name"])
            broker = Broker.objects.filter(company=company, name=broker_name).first()
            if broker is None:
                raise _RowError(f"Broker '{broker_name}' is not in the system.", ["broker_name"])
            clean["broker"] = broker

        elif source == SourceCode.REFERRAL:
            if not row.get("referrer_name"):
                raise _RowError("Referrer name is required for referral leads.", ["referrer_name"])
            clean["referrer_name"] = row["referrer_name"]

        # Campaign attribution is required for any row that names a campaign or a
        # channel (campaign is mandatory for every campaign type, per spec).
        BulkLeadImportService._resolve_campaign(company, source, row, clean)
        return clean

    @staticmethod
    def _resolve_campaign(company, source, row, clean):
        filled = [(col, row[col]) for col in _CHANNEL_COLUMNS if row.get(col)]
        campaign_name = row.get("campaign_name", "")

        if source == SourceCode.CAMPAIGN:
            if not campaign_name:
                raise _RowError("Campaign name is required for campaign leads.", ["campaign_name"])
            if len(filled) != 1:
                raise _RowError(
                    "Fill exactly one channel column "
                    "(event / tv_ad / street_ad / social_media_ad / exhibition).",
                    _CHANNEL_COLS)
        elif not campaign_name and not filled:
            return  # call_center / referral without campaign attribution: fine.
        elif not campaign_name:
            raise _RowError("Campaign name is required when a channel is filled.", ["campaign_name"])

        campaign = Campaign.objects.filter(company=company, name=campaign_name).first()
        if campaign is None:
            raise _RowError(f"Campaign '{campaign_name}' is not in the system.", ["campaign_name"])

        if not filled:
            raise _RowError("Select a channel column for this campaign lead.", _CHANNEL_COLS)
        col, rec_name = filled[0]
        ctype, model = _CHANNEL_COLUMNS[col]
        rec = model.objects.filter(campaign=campaign, name=rec_name).first()
        if rec is None:
            raise _RowError(
                f"{col} '{rec_name}' was not found in campaign '{campaign_name}'.", [col])
        clean.update(
            campaign=campaign, campaign_child_type=ctype,
            campaign_child_id=str(rec.id),
            attribution_event=rec if ctype == ChannelType.EVENT else None,
        )

    @staticmethod
    def _create(*, company, actor, clean, auto, request_meta):
        source = clean["source"]
        origin = Origin.BROKER if source == SourceCode.BROKER else Origin.DIRECT
        # Broker leads never auto-distribute (§8.5); everything else follows policy.
        do_auto = auto and source != SourceCode.BROKER
        LeadCreationService.create(
            company=company, actor=actor, source_code=source,
            name=clean["name"], phone=clean["phone"],
            country_code=clean["country_code"], email=clean["email"],
            language=clean["language"], origin=origin,
            broker_owner=clean.get("broker"),
            referrer_name=clean.get("referrer_name", ""),
            campaign=clean.get("campaign"),
            campaign_child_type=clean.get("campaign_child_type", ""),
            campaign_child_id=clean.get("campaign_child_id"),
            attribution_event=clean.get("attribution_event"),
            metadata={"notes": clean["notes"]} if clean["notes"] else None,
            auto_distribute=do_auto, notify=do_auto,
            request_meta=request_meta,
        )


class _RowError(Exception):
    """A rejected row plus the column(s) whose value caused the rejection, so the
    export can highlight only those cells (task: per-cell red, not whole row)."""

    def __init__(self, message, fields=None):
        super().__init__(message)
        self.fields = list(fields or [])


# Channel columns share a rejection (only one may be filled / must match).
_CHANNEL_COLS = list(_CHANNEL_COLUMNS.keys())


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _normalize_phone(phone: str) -> str:
    return "".join(ch for ch in (phone or "") if ch.isdigit())


def _resolve_language(lang_code: str, company):
    lang_code = (lang_code or "").strip()
    if not lang_code:
        from apps.policies.services import PolicyResolver
        from apps.policies.constants import PolicyCode
        lang_code = PolicyResolver.option_code(company, PolicyCode.LANGUAGE_DEFAULT, default="ar")
    language = Language.objects.filter(code=lang_code, is_active=True).first()
    if language is None:
        raise _RowError(f"Language '{lang_code}' is not in the system.", ["language"])
    return language
